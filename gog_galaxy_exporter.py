import os
import sys
import sqlite3
import json
import csv
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class GOGGalaxyExporter:
    def __init__(self):
        self.db_path = None
        self.connection = None
        
    def find_database_path(self) -> Optional[str]:
        """Find GOG Galaxy database path on Windows and macOS"""
        system = sys.platform.lower()
        
        if system.startswith("win"):  # Handles win32, win64, windows, etc.
            # Common Windows paths for GOG Galaxy
            possible_paths = [
                os.path.expandvars(r"%PROGRAMDATA%\GOG.com\Galaxy\storage\galaxy-2.0.db"),
                os.path.expandvars(r"%LOCALAPPDATA%\GOG.com\Galaxy\storage\galaxy-2.0.db"),
                os.path.expanduser(r"~\AppData\Local\GOG.com\Galaxy\storage\galaxy-2.0.db"),
                os.path.expanduser(r"~\AppData\Roaming\GOG.com\Galaxy\storage\galaxy-2.0.db"),
            ]
        elif system == "darwin":  # macOS
            possible_paths = [
                os.path.expanduser("~/Library/Application Support/GOG.com/Galaxy/storage/galaxy-2.0.db"),
                os.path.expanduser("~/Library/Preferences/GOG.com/Galaxy/storage/galaxy-2.0.db"),
            ]
        else:
            print(f"Unsupported operating system: {system}")
            return None
            
        for path in possible_paths:
            if os.path.exists(path):
                print(f"Found GOG Galaxy database at: {path}")
                return path
                
        print("GOG Galaxy database not found in standard locations.")
        return None
    
    def connect_database(self) -> bool:
        """Connect to GOG Galaxy database in read-only shared mode"""
        self.db_path = self.find_database_path()
        if not self.db_path:
            return False
            
        try:
            # Open in read-only mode with shared cache
            connection_string = f"file:{self.db_path}?mode=ro"
            self.connection = sqlite3.connect(connection_string, uri=True)
            self.connection.row_factory = sqlite3.Row
            print("Successfully connected to GOG Galaxy database (read-only mode)")
            return True
        except sqlite3.Error as e:
            print(f"Error connecting to database: {e}")
            return False
    
    def get_games_data(self) -> List[Dict[str, Any]]:
        db_path = self.find_database_path()
        if not db_path:
            return []
            
        try:
            connection_string = f"file:{db_path}?mode=ro"
            conn = sqlite3.connect(connection_string, uri=True)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            # First, get the current user ID
            cursor.execute("SELECT id FROM Users LIMIT 1")
            user_row = cursor.fetchone()
            if not user_row:
                print("No user found in database")
                return []
            
            user_id = user_row['id']
            
            # Get GamePieceTypes mapping for dynamic lookup
            cursor.execute("SELECT id, type FROM GamePieceTypes")
            piece_types = cursor.fetchall()
            piece_type_map = {row['type']: row['id'] for row in piece_types}
            
            print(f"Available GamePieceTypes: {list(piece_type_map.keys())}")
            
            # Enhanced query that includes GamePieces data
            query = """
            SELECT DISTINCT
                lr.releaseKey as release_key,
                ptr.gogId as product_id,
                ld.title,
                d.releaseDate as release_date,
                d.description,
                d.slug,
                CASE 
                    WHEN lr.releaseKey LIKE 'gog_%' THEN 'gog'
                    WHEN lr.releaseKey LIKE 'steam_%' THEN 'steam'
                    WHEN lr.releaseKey LIKE 'epic_%' THEN 'epic'
                    WHEN lr.releaseKey LIKE 'xboxone_%' THEN 'xboxone'
                    WHEN lr.releaseKey LIKE 'amazon_%' THEN 'amazon'
                    ELSE SUBSTR(lr.releaseKey, 1, INSTR(lr.releaseKey, '_') - 1)
                END as platform_name,
                gt.minutesInGame,
                lpd.lastPlayedDate as last_played,
                ppd.purchaseDate,
                ppd.addedDate,
                rp.gameId,
                rp.isDlc,
                rp.isVisibleInLibrary,
                lr.id as library_id
            FROM LibraryReleases lr
            LEFT JOIN ProductsToReleaseKeys ptr ON lr.releaseKey = ptr.releaseKey
            LEFT JOIN LimitedDetails ld ON ptr.gogId = ld.productId
            LEFT JOIN Details d ON ld.id = d.limitedDetailsId
            LEFT JOIN ProductPurchaseDates ppd ON lr.releaseKey = ppd.gameReleaseKey AND ppd.userId = ?
            LEFT JOIN GameTimes gt ON lr.releaseKey = gt.releaseKey AND gt.userId = ?
            LEFT JOIN LastPlayedDates lpd ON lr.releaseKey = lpd.gameReleaseKey AND lpd.userId = ?
            LEFT JOIN ReleaseProperties rp ON lr.releaseKey = rp.releaseKey
            WHERE lr.userId = ?
                AND (rp.isVisibleInLibrary IS NULL OR rp.isVisibleInLibrary = 1)
            ORDER BY COALESCE(ld.title, lr.releaseKey)
            """
            
            cursor.execute(query, (user_id, user_id, user_id, user_id))
            rows = cursor.fetchall()
            
            # Get user tags for releases
            cursor.execute("""
                SELECT releaseKey, tag 
                FROM UserReleaseTags 
                WHERE userId = ?
            """, (user_id,))
            
            tag_rows = cursor.fetchall()
            tags_dict = {}
            for tag_row in tag_rows:
                release_key = tag_row['releaseKey']
                if release_key not in tags_dict:
                    tags_dict[release_key] = []
                tags_dict[release_key].append(tag_row['tag'])
            
            # Get GamePieces data for all relevant piece types
            relevant_types = [
                'title', 'originalTitle', 'sortingTitle', 'originalSortingTitle',
                'meta', 'originalMeta', 'summary',
                'reviewScore', 'myRating',
                'storeFeatures', 'storeOsCompatibility', 'storeMedia', 'storeImages',
                'osCompatibility', 'media', 'originalImages',
                'storeTags', 'changelog', 'goodies',
                'localizations', 'preferredLocalization'
            ]
            type_ids = [str(piece_type_map[t]) for t in relevant_types if t in piece_type_map]
            
            if type_ids:
                placeholders = ','.join(['?' for _ in type_ids])
                cursor.execute(f"""
                    SELECT 
                        gp.releaseKey,
                        gpt.type as piece_type,
                        gp.value
                    FROM GamePieces gp
                    JOIN GamePieceTypes gpt ON gp.gamePieceTypeId = gpt.id
                    WHERE gpt.id IN ({placeholders})
                """, type_ids)
                
                game_pieces_rows = cursor.fetchall()
                game_pieces_dict = {}
                for piece_row in game_pieces_rows:
                    release_key = piece_row['releaseKey']
                    piece_type = piece_row['piece_type']
                    if release_key not in game_pieces_dict:
                        game_pieces_dict[release_key] = {}
                    game_pieces_dict[release_key][piece_type] = piece_row['value']
            else:
                game_pieces_dict = {}
            
            # Convert rows to standardized game data structure
            games_data = []
            for row in rows:
                row_dict = dict(row)
                tags = tags_dict.get(row_dict['release_key'], [])
                game_pieces = game_pieces_dict.get(row_dict['release_key'], {})
                standardized_game = self._standardize_game_data(row_dict, tags, game_pieces)
                games_data.append(standardized_game)
            
            conn.close()
            print(f"Found {len(games_data)} games")
            return games_data
                
        except sqlite3.Error as e:
            print(f"Database error: {e}")
            return []
    
    @staticmethod
    def _standardize_game_data(row_dict: Dict[str, Any], tags: List[str], game_pieces: Dict[str, Any]) -> Dict[str, Any]:
        """Standardize game data structure to match export expectations"""
        # Extract data from GamePieces JSON
        review_score = None
        my_rating = None
        critics_score = None
        features = []
        os_compatibility = []
        screenshots = []
        store_tags = []
        developers = []
        publishers = []
        genres = []
        themes = []
        original_release_date = None
        
        # Enhanced title extraction with priority order
        game_title = None
        title_sources = ['title', 'originalTitle', 'sortingTitle', 'originalSortingTitle']
        
        # First try database title
        if row_dict.get('title'):
            game_title = row_dict.get('title')
        else:
            # Try GamePieces titles in order of preference
            for title_type in title_sources:
                if title_type in game_pieces:
                    try:
                        title_data = json.loads(game_pieces[title_type])
                        if isinstance(title_data, dict):
                            game_title = title_data.get('title') or title_data.get('name')
                        elif isinstance(title_data, str):
                            game_title = title_data
                        if game_title:
                            break
                    except (json.JSONDecodeError, KeyError):
                        continue
        
        # Fallback to parsing release key if still no title
        if not game_title:
            release_key = row_dict.get('release_key', '')
            if '_' in release_key:
                game_title = release_key.split('_', 1)[1]
            else:
                game_title = release_key
        
        # Extract review scores
        if 'reviewScore' in game_pieces:
            try:
                score_data = json.loads(game_pieces['reviewScore'])
                review_score = score_data.get('score', {}).get('value')
            except (json.JSONDecodeError, KeyError):
                pass
                
        if 'myRating' in game_pieces:
            try:
                rating_data = json.loads(game_pieces['myRating'])
                my_rating = rating_data.get('rating')
            except (json.JSONDecodeError, KeyError):
                pass
        
        # Extract originalMeta data
        if 'originalMeta' in game_pieces:
            try:
                meta_data = json.loads(game_pieces['originalMeta'])
                critics_score = meta_data.get('criticsScore')
                developers = meta_data.get('developers', [])
                publishers = meta_data.get('publishers', [])
                genres = meta_data.get('genres', [])
                themes = meta_data.get('themes', [])
                
                # Handle release date (Unix timestamp)
                release_timestamp = meta_data.get('releaseDate')
                if release_timestamp:
                    try:
                        from datetime import datetime
                        original_release_date = datetime.fromtimestamp(release_timestamp).strftime('%Y-%m-%d')
                    except (ValueError, OSError):
                        pass
            except (json.JSONDecodeError, KeyError):
                pass
        
        # Also try to extract from 'meta' GamePiece (id 394) as fallback
        if not original_release_date and 'meta' in game_pieces:
            try:
                meta_data = json.loads(game_pieces['meta'])
                release_timestamp = meta_data.get('releaseDate')
                if release_timestamp:
                    try:
                        from datetime import datetime
                        original_release_date = datetime.fromtimestamp(release_timestamp).strftime('%Y-%m-%d')
                    except (ValueError, OSError):
                        pass
            except (json.JSONDecodeError, KeyError):
                pass
        
        # Extract description from summary (GamePieceType 401)
        description = row_dict.get('description', '')
        if 'summary' in game_pieces:
            try:
                summary_data = json.loads(game_pieces['summary'])
                if isinstance(summary_data, dict):
                    description = summary_data.get('description') or summary_data.get('summary', '')
                elif isinstance(summary_data, str):
                    description = summary_data
            except (json.JSONDecodeError, KeyError):
                pass
                
        # Extract features
        if 'storeFeatures' in game_pieces:
            try:
                features_data = json.loads(game_pieces['storeFeatures'])
                features = [f.get('name', '') for f in features_data.get('features', [])]
            except (json.JSONDecodeError, KeyError):
                pass
                
        # Extract OS compatibility
        for os_type in ['storeOsCompatibility', 'osCompatibility']:
            if os_type in game_pieces:
                try:
                    os_data = json.loads(game_pieces[os_type])
                    os_compatibility = [os.get('name', '') for os in os_data.get('supported', [])]
                    if os_compatibility:
                        break
                except (json.JSONDecodeError, KeyError):
                    pass
                    
        # Extract media/screenshots
        for media_type in ['storeMedia', 'media']:
            if media_type in game_pieces:
                try:
                    media_data = json.loads(game_pieces[media_type])
                    screenshots = media_data.get('screenshots', [])
                    if screenshots:
                        break
                except (json.JSONDecodeError, KeyError):
                    pass
                    
        # Extract store tags
        if 'storeTags' in game_pieces:
            try:
                tags_data = json.loads(game_pieces['storeTags'])
                store_tags = [tag.get('name', '') for tag in tags_data.get('tags', [])]
            except (json.JSONDecodeError, KeyError):
                pass
        
        return {
            # Primary identifiers
            'release_key': row_dict.get('release_key'),
            'game_id': row_dict.get('gameId') or row_dict.get('release_key'),
            
            # Game info
            'title': game_title,
            'description': description,
            
            # Platform and visibility
            'platform': row_dict.get('platform_name'),
            'is_dlc': bool(row_dict.get('isDlc', 0)),
            'is_visible': bool(row_dict.get('isVisibleInLibrary', 1)),
            
            # Dates
            'release_date': original_release_date,
            'purchase_date': row_dict.get('purchaseDate'),
            'added_date': row_dict.get('addedDate'),
            'last_played': row_dict.get('last_played'),
            
            # Time tracking
            'time_played': row_dict.get('minutesInGame', 0),
            'playtime_hours': round((row_dict.get('minutesInGame', 0) or 0) / 60.0, 2),
            
            # Enhanced data from GamePieces
            'review_score': review_score,
            'my_rating': my_rating,
            'critics_score': critics_score,
            'features': features,
            'os_compatibility': os_compatibility,
            'screenshots_count': len(screenshots),
            'store_tags': store_tags,
            
            # Enhanced metadata from originalMeta
            'developers': developers,
            'publishers': publishers,
            'genres': genres,
            'themes': themes,
            
            # Additional data
            'tags': tags,
            'library_id': row_dict.get('library_id')
        }
    
    def export_to_json(self, games_data: List[Dict], filename: str | None = None) -> str:
        """Export games data to JSON file"""
        if not filename:
            filename = f"gog_galaxy_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        filepath = Path(os.getcwd()) / filename
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(games_data, f, indent=2, ensure_ascii=False)
        
        print(f"Data exported to JSON: {filepath}")
        return str(filepath)
    
    def export_to_csv(self, games_data: List[Dict], filename: str | None = None) -> str:
        """Export games data to CSV file"""
        if not filename:
            filename = f"gog_galaxy_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        filepath = Path(os.getcwd()) / filename
        
        if not games_data:
            print("No data to export")
            return str(filepath)
        
        # Use the standardized data structure directly with enhanced fields
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            if games_data:
                # Define the CSV field order including enhanced GamePieces data
                fieldnames = [
                    'release_key', 'game_id', 'title', 'description', 
                    'platform', 'is_dlc', 'is_visible', 'release_date', 
                    'purchase_date', 'added_date', 'last_played', 
                    'time_played', 'playtime_hours', 'review_score', 'my_rating', 
                    'critics_score', 'features', 'os_compatibility', 'screenshots_count', 
                    'store_tags', 'developers', 'publishers', 'genres', 'themes', 
                    'tags', 'library_id'
                ]
                
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                
                for game in games_data:
                    # Convert list fields to comma-separated strings for CSV
                    csv_game = game.copy()
                    for field in ['features', 'os_compatibility', 'store_tags', 'developers', 'publishers', 'genres', 'themes', 'tags']:
                        if field in csv_game and isinstance(csv_game[field], list):
                            csv_game[field] = ', '.join(str(item) for item in csv_game[field] if item)
                    
                    writer.writerow(csv_game)
        
        print(f"Data exported to CSV: {filepath}")
        return str(filepath)
    
    def export_to_xlsx(self, games_data: List[Dict[str, Any]]) -> str:
        """Export games data to Excel (.xlsx) format with table formatting"""
        if not OPENPYXL_AVAILABLE:
            print("Warning: openpyxl not available. Install with: pip install openpyxl")
            return ""
            
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"gog_galaxy_export_{timestamp}.xlsx"
        
        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            
            if ws is None:
                print("Error: Failed to create worksheet")
                return ""
                
            ws.title = "GOG Galaxy Games"
            
            if not games_data:
                print("No games data to export")
                return ""
            
            # Define headers based on the first game's keys
            headers = list(games_data[0].keys())
            
            # Write headers
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data rows
            for row_idx, game in enumerate(games_data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = game.get(header, '')
                    
                    # Handle list fields by converting to comma-separated strings
                    if isinstance(value, list):
                        value = ', '.join(str(item) for item in value if item)
                    elif value is None:
                        value = ''
                    
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Create table
            if len(games_data) > 0:
                # Define table range
                end_column = chr(ord('A') + len(headers) - 1)
                if len(headers) > 26:
                    # Handle columns beyond Z (AA, AB, etc.)
                    first_letter = chr(ord('A') + (len(headers) - 1) // 26 - 1)
                    second_letter = chr(ord('A') + (len(headers) - 1) % 26)
                    end_column = first_letter + second_letter
                
                table_range = f"A1:{end_column}{len(games_data) + 1}"
                
                # Create and configure table
                table = Table(displayName="GOGGamesTable", ref=table_range)
                
                # Add table style
                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                table.tableStyleInfo = style
                
                # Add table to worksheet
                ws.add_table(table)
            
            # Auto-adjust column widths
            for col_idx, column in enumerate(ws.columns, 1):
                max_length = 0
                
                # Get column letter from index instead of cell attribute
                if col_idx <= 26:
                    column_letter = chr(ord('A') + col_idx - 1)
                else:
                    # Handle columns beyond Z (AA, AB, etc.)
                    first_letter = chr(ord('A') + (col_idx - 1) // 26 - 1)
                    second_letter = chr(ord('A') + (col_idx - 1) % 26)
                    column_letter = first_letter + second_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width with some padding, max 50 characters
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            try:
                wb.save(filename)
                file_size = os.path.getsize(filename)
                print(f"Excel export saved to: {filename}")
                print(f"File size: {file_size:,} bytes")
                print(f"Games exported: {len(games_data)}")
                return filename
            except Exception as e:
                print(f"Error saving Excel file: {e}")
                return ""
        except Exception as e:
            print(f"Error creating Excel workbook: {e}")
            return ""
    
    def close_connection(self):
        """Close database connection"""
        if self.connection:
            self.connection.close()
            self.connection = None
    
    def export_all_data(self, export_format: str = 'both') -> bool:
        games_data = self.get_games_data()
        
        if not games_data:
            print("No games data found to export.")
            return False
        
        try:
            success = True
            
            if export_format.lower() in ['json', 'both', 'all']:
                self.export_to_json(games_data)
            
            if export_format.lower() in ['csv', 'both', 'all']:
                self.export_to_csv(games_data)
                
            if export_format.lower() in ['xlsx', 'excel', 'all']:
                if OPENPYXL_AVAILABLE:
                    self.export_to_xlsx(games_data)
                else:
                    print("Excel export not available. Install openpyxl: pip install openpyxl")
                    success = False
            
            return success
            
        finally:
            self.close_connection()


def main():
    """Main function"""
    print("GOG Galaxy Data Exporter")
    print("=" * 40)
    
    # Check if openpyxl is available and inform user
    if not OPENPYXL_AVAILABLE:
        print("Note: Excel export not available. Install with: pip install openpyxl")
        print()
    
    exporter = GOGGalaxyExporter()
    
    # Check command line arguments
    export_format = 'both'
    if len(sys.argv) > 1:
        format_arg = sys.argv[1].lower()
        if format_arg in ['json', 'csv', 'both', 'xlsx', 'excel', 'all']:
            export_format = format_arg
        else:
            print("Usage: python gog_galaxy_exporter.py [json|csv|xlsx|excel|both|all]")
            print("  json  - Export to JSON format only")
            print("  csv   - Export to CSV format only")
            print("  xlsx  - Export to Excel format only")
            print("  excel - Export to Excel format only")
            print("  both  - Export to both JSON and CSV (default)")
            print("  all   - Export to JSON, CSV, and Excel")
            sys.exit(1)
    
    success = exporter.export_all_data(export_format)
    
    if success:
        print("\nExport completed successfully!")
    else:
        print("\nExport failed. Please ensure GOG Galaxy is installed and has been run at least once.")
        sys.exit(1)


if __name__ == "__main__":
    main()
